import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import datetime
import re

######### YOUR FILENAME HERE ###############
file_path = 'Applications_Fall.xlsx'
############################################

workbook = load_workbook(file_path)
sheet = workbook.active


def get_job_description(job_link):
    response = requests.get(job_link)
    soup = BeautifulSoup(response.content, 'html.parser')

    description_divs = soup.find_all('div', attrs={'class': re.compile(r'description', re.I), 'id': re.compile(r'description', re.I)})
    description = "\n".join([div.get_text(strip=True) for div in description_divs])

    if not description:
        full_text = soup.get_text(separator=' ', strip=True)
        description = full_text

    # # Further split the text if "Equal Opportunity Employer" is found
    # if "Equal Opportunity Employer" in description:
    #     description = description.split("Equal Opportunity Employer")[0]

    return description.strip() if description else "No description found"


def get_platform(job_link, company):
    jl = job_link.lower()
    if "linkedin"  in job_link: return "LinkedIn"
    elif "greenhouse" in job_link: return "Greenhouse"
    elif "lever" in job_link: return "Lever"
    elif "indeed" in job_link: return "Indeed"
    elif "glassdoor" in job_link: return "Glassdoor"
    elif "workday" in job_link: return "Workday"
    elif company and re.search(rf"//(?:www\.)?{re.escape(company.lower())}\.", job_link):
        return f"{company} Website"
    else: return "Other"

def get_job_role(job_link):

    response = requests.get(job_link)
    soup = BeautifulSoup(response.content, 'html.parser')
    job_role_tag = soup.find('h1') or soup.find('h2') or soup.find('div', {'class': re.compile(r'job-title', re.I)}) or soup.find('div', {'class': re.compile(r'jobPostingHeader', re.I)})
    if job_role_tag: return job_role_tag.get_text(strip=True)
    else: return input("Job role: ")


def add_job_application(job_link, referral):
    job_description = get_job_description(job_link)
    date_applied = datetime.now().strftime('%Y-%m-%d')
    platform = get_platform(job_link, company)
    position = get_job_role(job_link)

    next_row = sheet.max_row + 1

    sheet[f'F{next_row}'] = job_link       
    sheet[f'E{next_row}'] = job_description 
    sheet[f'D{next_row}'] = date_applied     #
    sheet[f'G{next_row}'] = referral
    sheet[f'B{next_row}'] = company
    sheet[f'H{next_row}'] = platform
    sheet[f'A{next_row}'] = position

    workbook.save(file_path)
    print(f"Data saved for {job_link}")


job_link = input("Enter the job link: ")
company = input("Enter the company: ")
referral = input("Referred? ")
add_job_application(job_link, referral)
